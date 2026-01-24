import sys
import os
import json
import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font

# === НАСТРОЙКИ ПУТЕЙ ===
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(current_dir))
sys.path.insert(0, project_root)
os.chdir(project_root)

from src.config import (
    SELECTED_ROUTE, SELECTED_MONTH, SELECTED_YEAR,
    SIMULATION_RESULT_FILE, OUTPUTS_DIR, SCHEDULE_BOOK_REPORT_FILE
)
from src.logger import get_logger

# Инициализируем логгер для этого модуля
logger = get_logger(__name__)

# OUTPUT_FILE = os.path.join(OUTPUTS_DIR, f"schedule_book_{SELECTED_ROUTE}_{SELECTED_MONTH}.xlsx")

MONTH_TO_NUM = {
    "Январь": "01",
    "Февраль": "02",
    "Март": "03",
    "Апрель": "04",
    "Май": "05",
    "Июнь": "06",
    "Июль": "07",
    "Август": "08",
    "Сентябрь": "09",
    "Октябрь": "10",
    "Ноябрь": "11",
    "Декабрь": "12",
}


def main():
    logger.info("Начинаем генерацию журнала нарядов")
    logger.debug(f"Маршрут: {SELECTED_ROUTE}, Месяц: {SELECTED_MONTH}, Год: {SELECTED_YEAR}")

    if not os.path.exists(SIMULATION_RESULT_FILE):
        logger.error(f"Файл симуляции не найден: {SIMULATION_RESULT_FILE}")
        logger.info("Для создания файла симуляции запустите run_simulation.py")
        return

    try:
        with open(SIMULATION_RESULT_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        logger.debug(f"Загружен файл симуляции с {len(data)} записями")
    except Exception as e:
        logger.error(f"Ошибка при чтении файла симуляции: {SIMULATION_RESULT_FILE}", exc_info=True)
        return

    # Сортируем дни (1, 2, 3...)
    days_sorted = sorted([k for k in data.keys() if k.isdigit()], key=int)
    logger.debug(f"Найдено дней для обработки: {len(days_sorted)}")

    rows = []

    logger.info("Формирование строк отчета...")
    for day_str in days_sorted:
        day_res = data[day_str]
        roster = day_res.get("roster", [])

        # Сортируем вагоны по номеру (как числа)
        try:
            roster.sort(key=lambda x: int(x.get("tram_number", 0)))
        except Exception as e:
            logger.warning(f"Не удалось отсортировать вагоны для дня {day_str}: {e}")

        for tram in roster:
            t_num = tram.get("tram_number", "?")

            # Данные 1 смены
            s1 = tram.get("shift_1") or {}
            d1 = s1.get("driver_name", "")
            t1 = s1.get("time_range", "")
            w1 = "(!)" if s1.get("warnings") else ""

            # Данные 2 смены
            s2 = tram.get("shift_2") or {}
            d2 = s2.get("driver_name", "")
            t2 = s2.get("time_range", "")
            w2 = "(!)" if s2.get("warnings") else ""

            day_dd = f"{int(day_str):02d}"  # 1 → 01
            month_mm = MONTH_TO_NUM.get(SELECTED_MONTH, "01")

            # Формируем строку
            row = {
                "Дата": f"{day_dd}.{month_mm}.{SELECTED_YEAR}",
                "Вагон": int(t_num) if str(t_num).isdigit() else t_num,

                "I Смена (Водитель)": f"{d1} {w1}".strip(),
                "I Время": t1,

                "II Смена (Водитель)": f"{d2} {w2}".strip(),
                "II Время": t2,

                "Проблемы": ", ".join(tram.get("issues", []))
            }
            rows.append(row)

    # Создаем DataFrame
    df = pd.DataFrame(rows)
    logger.debug(f"Создан DataFrame с {len(df)} строками")

    # Сохраняем
    try:
        os.makedirs(os.path.dirname(SCHEDULE_BOOK_REPORT_FILE), exist_ok=True)
        writer = pd.ExcelWriter(SCHEDULE_BOOK_REPORT_FILE, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name="Журнал")
        logger.debug("DataFrame записан в Excel файл")
    except Exception as e:
        logger.error(f"Ошибка при создании Excel файла: {SCHEDULE_BOOK_REPORT_FILE}", exc_info=True)
        return

    # Оформление
    ws = writer.book.active

    # Определяем строки, где заканчивается каждый день
    day_end_rows = []
    current_day = None

    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        cell_day = row[0].value  # колонка "Дата"
        if current_day is None:
            current_day = cell_day
        elif cell_day != current_day:
            day_end_rows.append(idx - 1)
            current_day = cell_day

    # последний день
    day_end_rows.append(ws.max_row)

    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    fill_warn = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Желтый если нарушение

    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    thick_bottom = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thick')
    )

    # Настраиваем заголовки
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Проходим по данным
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            row_idx = cell.row
            if row_idx in day_end_rows:
                cell.border = thick_bottom
            else:
                cell.border = border

            cell.alignment = Alignment(horizontal='center', vertical='center')

            val = str(cell.value)

            # Подсветка проблем
            if "(!)" in val:
                cell.fill = fill_warn

    # Ширина колонок
    ws.column_dimensions['A'].width = 15  # Дата
    ws.column_dimensions['B'].width = 8   # Вагон
    ws.column_dimensions['C'].width = 20  # Водитель 1
    ws.column_dimensions['D'].width = 15  # Время 1
    ws.column_dimensions['E'].width = 20  # Водитель 2
    ws.column_dimensions['F'].width = 15  # Время 2
    ws.column_dimensions['G'].width = 30  # Проблемы


    writer.close()
    logger.info(f"Журнал нарядов успешно создан: {SCHEDULE_BOOK_REPORT_FILE}")


if __name__ == "__main__":
    main()