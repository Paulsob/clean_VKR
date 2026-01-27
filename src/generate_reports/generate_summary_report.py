import sys
import os
import json
import glob
import re
import pandas as pd
from datetime import datetime, date
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# === НАСТРОЙКА ПУТЕЙ ===
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(current_dir))
sys.path.insert(0, project_root)
os.chdir(project_root)

from src.config import (
    SELECTED_MONTH, SELECTED_YEAR, SELECTED_ROUTE, PROCESS_ALL_ROUTES,
    SIMULATION_MODE  # <--- ДОБАВЛЕНО
)
from src.prepare_data.database import DataLoader
from src.logger import get_logger

logger = get_logger("SummaryReport")


# === ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ===

def prepare_absences_map(db_absences):
    absences_map = {}
    for item in db_absences:
        did = str(item["driver_id"])
        if did not in absences_map: absences_map[did] = {}
        current = item["from"]
        while current <= item["to"]:
            absences_map[did][current] = item["type"]
            current += pd.Timedelta(days=1).to_pytimedelta()
    return absences_map


def prepare_schedule_counts(db_schedules):
    counts = {}
    for schedule in db_schedules:
        if hasattr(schedule, 'model_dump'):
            s_dict = schedule.model_dump()
        elif hasattr(schedule, 'dict'):
            s_dict = schedule.dict()
        else:
            s_dict = schedule.__dict__

        r_num = str(s_dict.get("route_number") or s_dict.get("маршрут"))
        day_type = s_dict.get("day_type") or s_dict.get("день")
        trams = s_dict.get("trams") or s_dict.get("трамваи", [])

        shifts = 0
        for tram in trams:
            if hasattr(tram, 'model_dump'):
                t_dict = tram.model_dump()
            elif hasattr(tram, 'dict'):
                t_dict = tram.dict()
            elif isinstance(tram, dict):
                t_dict = tram
            else:
                t_dict = tram.__dict__

            if t_dict.get("смена_1") or t_dict.get("shift_1"): shifts += 1
            if t_dict.get("смена_2") or t_dict.get("shift_2"): shifts += 1

        if r_num not in counts: counts[r_num] = {'workday': 0, 'weekend': 0}

        if day_type == "рабочий":
            counts[r_num]['workday'] = shifts
        else:
            counts[r_num]['weekend'] = shifts
    return counts


def get_day_of_week_row(year, month, all_days):
    days_map = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    row = {"Маршрут": "", "Таб.№": "", "График": "", "Смен": "", "Часов": "", "Резерв (дн)": "День недели"}
    for day in all_days:
        try:
            dt = date(year, month, day)
            row[day] = days_map[dt.weekday()]
        except:
            row[day] = "?"
    return row


def style_worksheet(ws, all_days_len):
    """Стилизация + Автоширина + Запрет переноса"""
    fill_work = PatternFill("solid", fgColor="C6EFCE")
    fill_work_warn = PatternFill("solid", fgColor="FFEB9C")
    fill_reserve = PatternFill("solid", fgColor="FFC7CE")
    fill_rest = PatternFill("solid", fgColor="F2F2F2")
    fill_sick = PatternFill("solid", fgColor="FFFF00")
    fill_header = PatternFill("solid", fgColor="4472C4")
    fill_stat = PatternFill("solid", fgColor="D9E1F2")
    fill_guest_header = PatternFill("solid", fgColor="70AD47")
    fill_any_header = PatternFill("solid", fgColor="FFD966")  # Желтоватый для ANY

    font_header = Font(bold=True, color="FFFFFF")
    font_bold = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        first_val = str(row[0].value) if row[0].value else ""

        is_header_group = first_val in ["РАБОЧЕЕ ЯДРО", "ИЗБЫТОЧНЫЙ РЕЗЕРВ"]
        is_guest_header = "ПРИВЛЕЧЕННЫЕ" in first_val or "РЕЗЕРВ (НЕ ЗАКРЕПЛЕННЫЕ)" in first_val
        is_any_header = "РЕЗЕРВ ANY" in first_val
        # Определяем блок статистики по ключевым словам из обоих режимов
        is_stat_block = any(k in first_val for k in [
            "Количество закрытых смен", "Количество незакрытых смен",
            "Количество смен по расписанию", "Всего водителей",
            "Количество водителей, вышедших", "Количество водителей, находящихся"
        ])
        is_dow_row = str(row[-1].value) in ["Пн", "Вт", "Сб", "Вс"] or "День недели" in str(row[5].value)

        if is_header_group:
            for cell in row:
                cell.fill = fill_header
                cell.font = font_header
            continue

        if is_guest_header:
            for cell in row:
                cell.fill = fill_guest_header
                cell.font = font_header
            continue

        if is_any_header:
            for cell in row:
                cell.fill = fill_any_header
                cell.font = font_bold
            continue

        if is_stat_block:
            for cell in row:
                cell.fill = fill_stat
                cell.font = font_bold
                cell.border = border
            continue

        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            val = str(cell.value) if cell.value else ""

            if is_dow_row:
                cell.font = font_bold
                if val in ["Сб", "Вс"]: cell.fill = PatternFill("solid", fgColor="E2EFDA")
                continue

            if "БОЛЬНИЧНЫЙ" in val or "ОТПУСК" in val:
                cell.fill = fill_sick
            elif "(!)" in val or "⚠️" in val:
                cell.fill = fill_work_warn
            elif "ч (" in val:
                cell.fill = fill_work
            elif "РЕЗЕРВ" in val:
                cell.fill = fill_reserve
            elif val in ["В", "B", "О", "Б"]:
                cell.fill = fill_rest

    # Автоподбор ширины
    for column_cells in ws.columns:
        length = 0
        for cell in column_cells:
            if cell.value:
                length = max(length, len(str(cell.value)))
        adjusted_width = (length + 2) * 1.05
        if adjusted_width > 50: adjusted_width = 50  # Чуть шире для длинных названий статистики
        if adjusted_width < 5: adjusted_width = 5
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width


# === ЛОГИКА ОБРАБОТКИ ===

def get_schedule_from_driver(driver_obj):
    sch_type = "?"
    if hasattr(driver_obj, 'schedule_pattern'):
        sch_type = driver_obj.schedule_pattern
    elif hasattr(driver_obj, 'schedule'):
        sch_type = driver_obj.schedule
    if sch_type == "?" or sch_type is None:
        d_dump = driver_obj.model_dump() if hasattr(driver_obj, 'model_dump') else driver_obj.__dict__
        sch_type = d_dump.get('schedule') or d_dump.get('schedule_pattern') or d_dump.get('mode') or "?"
    return sch_type


def process_route(route_number, sim_file_path, assigned_drivers, all_drivers_db, absences_map, schedule_counts,
                  month_num, all_days):
    """
    Возвращает расширенную статистику для режимов Real/Strict
    """
    try:
        with open(sim_file_path, "r", encoding="utf-8") as f:
            sim_data = json.load(f)
    except Exception:
        return [], [], [], {}, {}, {}, {}  # Возвращаем пустые структуры

    native_drivers_map = {str(d.id): d for d in assigned_drivers}
    report_data_native = {did: {'days': {}, 'work_count': 0, 'reserve_count': 0, 'hours': 0} for did in
                          native_drivers_map}
    guest_drivers_data = {}

    # === СБОР СТАТИСТИКИ (ИНИЦИАЛИЗАЦИЯ) ===
    # Нам нужно собирать данные для каждого дня
    daily_counters = {d: {
        'closed_shifts': 0,  # 1. Закрытые смены (всего)
        'plan_shifts': 0,  # 3. Смены по расписанию
        'drivers_native': 0,  # 4. Свои на маршруте
        'drivers_guest': 0,  # 5. Чужие на маршруте
        'drivers_weekend_work': 0,  # 6. Работа в выходной (свой или чужой)
        'drivers_on_rest': 0,  # 7. На выходном (только свои, т.к. чужие по определению работают)
        'reserve_true': 0  # 8. Не хватило наряда (резерв)
    } for d in all_days}

    route_plan = schedule_counts.get(str(route_number), {'workday': 0, 'weekend': 0})

    # 1. ФАКТ (Симуляция) - заполняем работу
    for day_str, day_res in sim_data.items():
        if not day_str.isdigit(): continue
        day = int(day_str)

        # Считаем план на этот день
        dt = date(SELECTED_YEAR, month_num, day)
        is_weekend = dt.weekday() >= 5
        daily_counters[day]['plan_shifts'] = route_plan['weekend'] if is_weekend else route_plan['workday']

        for tram in day_res.get("roster", []):
            for shift_key in ["shift_1", "shift_2"]:
                s_info = tram.get(shift_key)
                if s_info and s_info.get("driver"):
                    daily_counters[day]['closed_shifts'] += 1  # +1 закрытая смена

                    raw_did = s_info["driver"].split(" ")[0]
                    wh = s_info.get("work_hours", 8.0)
                    rest = s_info.get("rest_before", 0)
                    rest_str = "—" if rest > 500 else f"{rest:.0f}"
                    val = f"{wh:.1f}ч (отд {rest_str})"
                    if s_info.get("warnings"): val += " (!)"

                    # Определяем объект водителя для проверки графика (работа в выходной)
                    current_driver_obj = native_drivers_map.get(raw_did)
                    if not current_driver_obj:
                        # Если это гость, ищем в общей базе
                        current_driver_obj = next((d for d in all_drivers_db if str(d.id) == raw_did), None)

                    # Проверка на работу в выходной (если по графику у него 'В')
                    if current_driver_obj:
                        status_in_plan = current_driver_obj.get_status_for_day(day)
                        # Если в плане Выходной, Отпуск или Больничный, но он работает -> это работа в выходной
                        if status_in_plan in ['В', 'B'] or status_in_plan.isdigit() == False:
                            daily_counters[day]['drivers_weekend_work'] += 1

                    if raw_did in report_data_native:
                        # СВОЙ
                        report_data_native[raw_did]['days'][day] = val
                        report_data_native[raw_did]['work_count'] += 1
                        report_data_native[raw_did]['hours'] += wh
                        daily_counters[day]['drivers_native'] += 1
                    else:
                        # ЧУЖОЙ
                        if raw_did not in guest_drivers_data:
                            if current_driver_obj:
                                guest_drivers_data[raw_did] = {'obj': current_driver_obj, 'days': {}, 'work_count': 0,
                                                               'hours': 0}
                        if raw_did in guest_drivers_data:
                            guest_drivers_data[raw_did]['days'][day] = val
                            guest_drivers_data[raw_did]['work_count'] += 1
                            guest_drivers_data[raw_did]['hours'] += wh
                        daily_counters[day]['drivers_guest'] += 1

    # 2. ЗАПОЛНЕНИЕ СТАТУСОВ (Только для СВОИХ) + Подсчет Резерва и Отдыхающих
    for did, d_obj in native_drivers_map.items():
        driver_absences = absences_map.get(did, {})
        for day in all_days:
            # Если водитель работал в этот день, мы его уже учли выше
            if day in report_data_native[did]['days']:
                continue

            check_date = date(SELECTED_YEAR, month_num, day)
            if check_date in driver_absences:
                a_type = driver_absences[check_date]
                report_data_native[did]['days'][day] = "БОЛЬНИЧНЫЙ" if a_type == "sick" else "ОТПУСК"
                # В 'strict' и 'real' не просили отдельно считать больных, но они попадают в "Общее кол-во"
                continue

            plan = d_obj.get_status_for_day(day)
            if plan in ["1", "2"]:
                report_data_native[did]['days'][day] = "РЕЗЕРВ"
                report_data_native[did]['reserve_count'] += 1
                daily_counters[day]['reserve_true'] += 1
            else:
                report_data_native[did]['days'][day] = plan
                # Если статус В, О, Б -> это отдыхающий (для статистики REAL)
                if plan in ["В", "B", "О", "Б"]:
                    daily_counters[day]['drivers_on_rest'] += 1

    # 3. ФОРМИРОВАНИЕ СТРОК (СВОИ)
    active_rows = []
    reserve_rows = []
    sorted_native = sorted(report_data_native.keys(), key=lambda x: int(x) if x.isdigit() else x)
    for did in sorted_native:
        data = report_data_native[did]
        d_obj = native_drivers_map[did]
        row = {
            "Маршрут": route_number, "Таб.№": did, "График": get_schedule_from_driver(d_obj),
            "Смен": data['work_count'], "Часов": round(data['hours'], 1), "Резерв (дн)": data['reserve_count']
        }
        for day in all_days: row[day] = data['days'].get(day, "")
        if data['work_count'] > 0:
            active_rows.append(row)
        else:
            reserve_rows.append(row)

    # 4. ГОСТИ
    guest_rows = []
    sorted_guests = sorted(guest_drivers_data.keys(), key=lambda x: int(x) if x.isdigit() else x)
    for did in sorted_guests:
        g_data = guest_drivers_data[did]
        d_obj = g_data['obj']
        row = {
            "Маршрут": route_number, "Таб.№": f"{did} (Рез)", "График": get_schedule_from_driver(d_obj),
            "Смен": g_data['work_count'], "Часов": round(g_data['hours'], 1), "Резерв (дн)": ""
        }
        for day in all_days: row[day] = g_data['days'].get(day, "")
        guest_rows.append(row)

    # 5. ДАННЫЕ ДЛЯ ГЛОБАЛЬНОГО СВОДА
    guest_raw_global = {}
    any_drivers_raw_global = {}

    for did, g_data in guest_drivers_data.items():
        d_obj = g_data['obj']
        info = {'obj': g_data['obj'], 'updates': {}}
        for day, val in g_data['days'].items():
            info['updates'][day] = f"{val} №{route_number}"

        if str(d_obj.assigned_route_number).upper() == "ANY":
            any_drivers_raw_global[did] = info
        else:
            guest_raw_global[did] = info

    # 6. СТАТИСТИКА (ЛОКАЛЬНАЯ ДЛЯ ЛИСТА МАРШРУТА - оставим упрощенную, как было, или сделаем детальную?)
    # Оставим упрощенную для листов маршрутов, чтобы не загромождать, а детальную вернем в daily_stats_raw

    # Для совместимости со старым кодом локального листа, сформируем базовые строки
    row_plan = {"Маршрут": "Количество смен по расписанию", "Таб.№": "", "График": "", "Смен": "", "Часов": "",
                "Резерв (дн)": ""}
    row_fact = {"Маршрут": "Количество водителей, вышедших на смену", "Таб.№": "", "График": "", "Смен": "",
                "Часов": "", "Резерв (дн)": ""}
    row_reserv = {"Маршрут": "Количество водителей без наряда (резерв)", "Таб.№": "", "График": "", "Смен": "",
                  "Часов": "", "Резерв (дн)": ""}

    for day in all_days:
        row_plan[day] = daily_counters[day]['plan_shifts']
        # В локальном отчете "Вышедшие" = Свои + Чужие
        row_fact[day] = daily_counters[day]['drivers_native'] + daily_counters[day]['drivers_guest']
        row_reserv[day] = daily_counters[day]['reserve_true']

    stats_block = [row_plan, row_fact, row_reserv]  # Локально можно проще

    # Возвращаем детальную статистику для глобального свода
    daily_stats_detailed = daily_counters
    daily_stats_detailed['total_drivers_native'] = len(native_drivers_map)  # Кол-во закрепленных

    return active_rows, reserve_rows, guest_rows, stats_block, daily_stats_detailed, guest_raw_global, any_drivers_raw_global


# === MAIN ===

def main():
    logger.info("Инициализация генератора отчетов...")
    db = DataLoader()
    db.load_all()

    absences_map = prepare_absences_map(db.absences)
    schedule_counts = prepare_schedule_counts(db.schedules)

    month_map = {"Январь": 1, "Февраль": 2, "Март": 3, "Апрель": 4, "Май": 5, "Июнь": 6,
                 "Июль": 7, "Август": 8, "Сентябрь": 9, "Октябрь": 10, "Ноябрь": 11, "Декабрь": 12}
    m_num = month_map.get(SELECTED_MONTH, 1)
    month_str_num = f"{m_num:02d}"
    results_dir = os.path.join("data", "results", f"{month_str_num}_{SELECTED_MONTH}_{SELECTED_YEAR}")
    search_pattern = os.path.join(results_dir, f"simulation_*_{SELECTED_MONTH}_{SELECTED_YEAR}.json")
    found_files = glob.glob(search_pattern)

    if not found_files:
        logger.error(f"Файлы симуляции не найдены в {results_dir}")
        return

    with open(found_files[0], 'r') as f:
        tmp = json.load(f)
        all_days = sorted([int(k) for k in tmp.keys() if k.isdigit()])

    target_route = str(SELECTED_ROUTE) if not PROCESS_ALL_ROUTES else None

    if PROCESS_ALL_ROUTES:
        filename_summary = f"summary_report_FULL_PARK_{SELECTED_MONTH}_{SELECTED_YEAR}.xlsx"
    else:
        filename_summary = f"summary_report_{SELECTED_ROUTE}_{SELECTED_MONTH}_{SELECTED_YEAR}.xlsx"

    directory_name = f"{month_str_num}_{SELECTED_MONTH}_{SELECTED_YEAR}"
    dynamic_summary_file = os.path.join("outputs", "SUMMARY_REPORTS", directory_name, filename_summary)

    os.makedirs(os.path.dirname(dynamic_summary_file), exist_ok=True)
    writer = pd.ExcelWriter(dynamic_summary_file, engine='openpyxl')

    # Глобальные контейнеры
    global_natives = []
    global_reserves_native = []
    global_unassigned_guests = {}
    global_any_drivers = {}

    # Сумматоры детальной статистики
    global_stats_detailed = {d: {
        'closed_shifts': 0, 'plan_shifts': 0,
        'drivers_native': 0, 'drivers_guest': 0,
        'drivers_weekend_work': 0, 'drivers_on_rest': 0,
        'reserve_true': 0
    } for d in all_days}
    global_total_drivers_count = 0

    dow_row = get_day_of_week_row(SELECTED_YEAR, m_num, all_days)

    def get_route_num(path):
        m = re.search(r"simulation_(\d+)_", os.path.basename(path))
        return int(m.group(1)) if m else 9999

    found_files.sort(key=get_route_num)
    processed_count = 0

    for file_path in found_files:
        route_num = str(get_route_num(file_path))
        if target_route and route_num != target_route: continue
        processed_count += 1

        route_assigned_drivers = [d for d in db.drivers if
                                  str(d.assigned_route_number) == route_num and d.month == SELECTED_MONTH]

        active, reserve, guests, stats, detailed_stats, guest_global_updates, any_drivers_updates = process_route(
            route_num, file_path, route_assigned_drivers, db.drivers, absences_map, schedule_counts, m_num, all_days
        )

        if PROCESS_ALL_ROUTES:
            global_natives.extend(active)
            global_reserves_native.extend(reserve)
            global_total_drivers_count += detailed_stats['total_drivers_native']

            # Суммируем дневную статистику
            for d in all_days:
                src = detailed_stats[d]
                dst = global_stats_detailed[d]
                dst['closed_shifts'] += src['closed_shifts']
                dst['plan_shifts'] += src['plan_shifts']
                dst['drivers_native'] += src['drivers_native']
                dst['drivers_guest'] += src['drivers_guest']
                dst['drivers_weekend_work'] += src['drivers_weekend_work']
                dst['drivers_on_rest'] += src['drivers_on_rest']
                dst['reserve_true'] += src['reserve_true']

            # Агрегация гостей
            for did, info in guest_global_updates.items():
                d_obj = info['obj']
                if not d_obj.assigned_route_number or d_obj.assigned_route_number == "0":
                    if did not in global_unassigned_guests:
                        global_unassigned_guests[did] = {'obj': d_obj, 'days': {}}
                    for day, val in info['updates'].items():
                        global_unassigned_guests[did]['days'][day] = val

            # Агрегация ANY
            for did, info in any_drivers_updates.items():
                d_obj = info['obj']
                if did not in global_any_drivers:
                    global_any_drivers[did] = {'obj': d_obj, 'days': {}}
                for day, val in info['updates'].items():
                    if day in global_any_drivers[did]['days']:
                        existing = global_any_drivers[did]['days'][day]
                        global_any_drivers[did]['days'][day] = f"КОНФЛИКТ: {existing} И {val}"
                    else:
                        global_any_drivers[did]['days'][day] = val

        # --- ЗАПИСЬ ЛИСТА МАРШРУТА ---
        sheet_rows = [dow_row]
        h_act = {"Таб.№": "РАБОЧЕЕ ЯДРО", "График": "", "Смен": "", "Часов": "", "Резерв (дн)": ""}
        for d in all_days: h_act[d] = ""
        sheet_rows.append(h_act)
        for r in active: c = r.copy(); del c["Маршрут"]; sheet_rows.append(c)

        sep = {k: "" for k in h_act}
        sheet_rows.append(sep)
        h_res = {"Таб.№": "ИЗБЫТОЧНЫЙ РЕЗЕРВ", "График": "", "Смен": "", "Часов": "", "Резерв (дн)": ""}
        for d in all_days: h_res[d] = ""
        sheet_rows.append(h_res)
        for r in reserve: c = r.copy(); del c["Маршрут"]; sheet_rows.append(c)
        sheet_rows.append(sep)

        # Локальная статистика (упрощенная)
        for sr in stats:
            sr_c = sr.copy();
            title = sr_c["Маршрут"];
            del sr_c["Маршрут"]
            sr_c["Таб.№"] = title
            sheet_rows.append(sr_c)

        if guests:
            sheet_rows.append(sep)
            h_guest = {"Таб.№": "ПРИВЛЕЧЕННЫЕ", "График": "", "Смен": "", "Часов": "", "Резерв (дн)": ""}
            for d in all_days: h_guest[d] = ""
            sheet_rows.append(h_guest)
            for r in guests: c = r.copy(); del c["Маршрут"]; sheet_rows.append(c)

        df = pd.DataFrame(sheet_rows)
        cols = ["Таб.№", "График", "Смен", "Часов", "Резерв (дн)"] + all_days
        df = df[cols]
        sn = f"Маршрут {route_num}"
        df.to_excel(writer, index=False, sheet_name=sn)
        style_worksheet(writer.sheets[sn], len(all_days))

    # --- ОБЩИЙ СВОД (С ИЗМЕНЕНИЯМИ) ---
    if PROCESS_ALL_ROUTES and processed_count > 0:
        logger.info("Формирование общего свода...")
        common_rows = [dow_row]
        sep = {k: "" for k in dow_row}

        # 1. Свои Активные
        h_glob = {"Маршрут": "РАБОЧЕЕ ЯДРО", "Таб.№": "", "График": "", "Смен": "", "Часов": "", "Резерв (дн)": ""}
        for d in all_days: h_glob[d] = ""
        common_rows.append(h_glob)
        common_rows.extend(global_natives)
        common_rows.append(sep)

        # 2. Свои Резерв
        h_glob_res = {"Маршрут": "ИЗБЫТОЧНЫЙ РЕЗЕРВ", "Таб.№": "", "График": "", "Смен": "", "Часов": "",
                      "Резерв (дн)": ""}
        for d in all_days: h_glob_res[d] = ""
        common_rows.append(h_glob_res)
        common_rows.extend(global_reserves_native)
        common_rows.append(sep)

        # 3. Резервные (незакрепленные) водители
        if global_unassigned_guests:
            h_unassigned = {"Маршрут": "РЕЗЕРВ (НЕ ЗАКРЕПЛЕННЫЕ)", "Таб.№": "", "График": "", "Смен": "", "Часов": "",
                            "Резерв (дн)": ""}
            for d in all_days: h_unassigned[d] = ""
            common_rows.append(h_unassigned)
            sorted_u = sorted(global_unassigned_guests.keys(), key=lambda x: int(x) if x.isdigit() else x)
            for did in sorted_u:
                info = global_unassigned_guests[did];
                d_obj = info['obj'];
                days_data = info['days']
                row = {"Маршрут": "Резерв", "Таб.№": did, "График": get_schedule_from_driver(d_obj),
                       "Смен": len(days_data), "Часов": "", "Резерв (дн)": ""}
                for day in all_days: row[day] = days_data.get(day, "")
                common_rows.append(row)
            common_rows.append(sep)

        # 4. Водители ANY (резервные)
        if global_any_drivers:
            h_any = {"Маршрут": "РЕЗЕРВ ANY (ВОДИТЕЛИ С МАРШРУТОМ ANY)", "Таб.№": "", "График": "", "Смен": "",
                     "Часов": "", "Резерв (дн)": ""}
            for d in all_days: h_any[d] = ""
            common_rows.append(h_any)
            sorted_any = sorted(global_any_drivers.keys(), key=lambda x: int(x) if x.isdigit() else x)
            for did in sorted_any:
                info = global_any_drivers[did];
                d_obj = info['obj'];
                days_data = info['days']
                total_hours = 0;
                work_days = 0;
                conflicts = 0;
                reserve_days = 0
                filled_days_data = {}
                for day in all_days:
                    if day in days_data:
                        day_val = days_data[day]
                        filled_days_data[day] = day_val
                        if "КОНФЛИКТ" in day_val:
                            conflicts += 1
                        else:
                            work_days += 1
                            match = re.search(r'(\d+\.?\d*)ч', day_val)
                            if match: total_hours += float(match.group(1))
                    else:
                        check_date = date(SELECTED_YEAR, m_num, day)
                        driver_absences = absences_map.get(did, {})
                        if check_date in driver_absences:
                            a_type = driver_absences[check_date]
                            filled_days_data[day] = "БОЛЬНИЧНЫЙ" if a_type == "sick" else "ОТПУСК"
                        else:
                            plan = d_obj.get_status_for_day(day)
                            if plan in ["1", "2"]:
                                filled_days_data[day] = "РЕЗЕРВ"
                                reserve_days += 1
                            else:
                                filled_days_data[day] = plan
                                # В режиме REAL здесь можно посчитать отдыхающих ANY, но пока не будем усложнять глобальную стату

                row = {
                    "Маршрут": "ANY", "Таб.№": f"{did}" + (" ⚠️" if conflicts > 0 else ""),
                    "График": get_schedule_from_driver(d_obj), "Смен": work_days,
                    "Часов": round(total_hours, 1) if total_hours > 0 else "",
                    "Резерв (дн)": reserve_days if reserve_days > 0 else (
                        f"Конфликтов: {conflicts}" if conflicts > 0 else "")
                }
                for day in all_days: row[day] = filled_days_data.get(day, "")
                common_rows.append(row)

            common_rows.append(sep)  # Отделяем статистику от ANY

        # 5. СТАТИСТИКА (В САМОМ НИЗУ)
        logger.info(f"Формирование блока статистики для режима: {SIMULATION_MODE}")

        # Подготовка данных
        stat_closed = {d: global_stats_detailed[d]['closed_shifts'] for d in all_days}
        stat_plan = {d: global_stats_detailed[d]['plan_shifts'] for d in all_days}
        # Незакрытые = План - Закрытые (но не меньше 0, если вдруг перевыполнение)
        stat_unclosed = {d: max(0, global_stats_detailed[d]['plan_shifts'] - global_stats_detailed[d]['closed_shifts'])
                         for d in all_days}

        stat_native_work = {d: global_stats_detailed[d]['drivers_native'] for d in all_days}
        stat_guest_work = {d: global_stats_detailed[d]['drivers_guest'] for d in all_days}

        stat_weekend_work = {d: global_stats_detailed[d]['drivers_weekend_work'] for d in all_days}
        stat_on_rest = {d: global_stats_detailed[d]['drivers_on_rest'] for d in all_days}
        stat_reserve = {d: global_stats_detailed[d]['reserve_true'] for d in all_days}

        # Создаем строки
        def create_stat_row(title, data_dict):
            r = {"Маршрут": title, "Таб.№": "", "График": "", "Смен": "", "Часов": "", "Резерв (дн)": ""}
            for d in all_days: r[d] = data_dict[d]
            return r

        row_1 = create_stat_row("Количество закрытых смен", stat_closed)
        row_2 = create_stat_row("Количество незакрытых смен", stat_unclosed)
        row_3 = create_stat_row("Количество смен по расписанию", stat_plan)
        row_4 = create_stat_row("Количество водителей, вышедших по расписанию на свой маршрут", stat_native_work)
        row_5 = create_stat_row("Количество водителей, вышедших по расписанию и не прикрепленных к маршруту",
                                stat_guest_work)
        row_6 = create_stat_row("Количество водителей, вышедших на смену в выходной день", stat_weekend_work)

        # Специфичные строки
        row_7_real = create_stat_row("Количество водителей, находящихся на выходном", stat_on_rest)
        row_8_real = create_stat_row("Количество водителей, которым не хватило наряда", stat_reserve)

        # Общее количество (константа для всех дней, либо динамически, если кол-во меняется)
        # Здесь берем общее число закрепленных
        row_total = {"Маршрут": "Всего водителей (закрепленных)", "Таб.№": global_total_drivers_count, "График": "",
                     "Смен": "", "Часов": "", "Резерв (дн)": ""}
        # Можно дублировать в ячейки дней, если нужно
        # for d in all_days: row_total[d] = global_total_drivers_count

        if SIMULATION_MODE == 'real':
            # REAL: 1-6, 7(rest), 8(reserve), 9(total)
            common_rows.extend([row_1, row_2, row_3, row_4, row_5, row_6, row_7_real, row_8_real, row_total])
        else:
            # STRICT: 1-6, 7(total). *Заметка: в промпте было 7 пунктов, но пропущен "резерв".
            # "Количество водителей, находящихся на выходном" (7) в Strict тоже просили.
            # "Общее количество" (8).
            # Исходя из текста "в режиме strict нужно... 6... 7. Общее количество", пропущены отдыхающие и резерв
            # Но по логике Strict ("строго с дырами"), переработки (п.6) быть не должно, но поле мы оставим (оно будет 0).
            # Сделаю ровно как в запросе:
            # 1. закрытые, 2. незакрытые, 3. по расписанию, 4. свои, 5. чужие, 6. выходной (зачем в стрикте?), 7. Общее
            common_rows.extend([row_1, row_2, row_3, row_4, row_5, row_6, row_total])

        df_glob = pd.DataFrame(common_rows)
        cols_glob = ["Маршрут", "Таб.№", "График", "Смен", "Часов", "Резерв (дн)"] + all_days
        df_glob = df_glob[cols_glob]

        df_glob.to_excel(writer, index=False, sheet_name="ОБЩИЙ СВОД")
        style_worksheet(writer.sheets["ОБЩИЙ СВОД"], len(all_days))

        wb = writer.book
        if "ОБЩИЙ СВОД" in wb.sheetnames:
            sheet = wb["ОБЩИЙ СВОД"]
            if sheet in wb._sheets:
                wb._sheets.remove(sheet)
            wb._sheets.insert(0, sheet)
            wb.active = 0

    writer.close()
    logger.info(f"Сводный отчет готов: {dynamic_summary_file}")


if __name__ == "__main__":
    main()