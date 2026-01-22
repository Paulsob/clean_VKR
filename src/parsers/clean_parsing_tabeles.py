import pandas as pd
import numpy as np
import os
import calendar
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ================= НАСТРОЙКИ =================
INPUT_FILE = '../../data/tabeles_2026/february_2026.xlsx'
OUTPUT_DIR = '../../data/tabeles_2026'

SHIFT_1 = 1
SHIFT_2 = 2
REST = 'В'

# Праздники 2026 (для 5х2)
HOLIDAYS_2026 = {
    (1, 1), (1, 2), (1, 3), (1, 4), (1, 5), (1, 6), (1, 7), (1, 8), (1, 9),
    (2, 23),
    (3, 8), (3, 9),
    (5, 1), (5, 9), (5, 11),
    (6, 12),
    (11, 4)
}

# Цикличные графики (кроме 5х2)
CYCLIC_PATTERNS = {
    '4x2': [1, 1, 1, 1, 0, 0],
    '3x2x3x1': [1, 1, 1, 0, 0, 1, 1, 1, 0],  # Ритм: 3р 2в 3р 1в
    '3x4': [1, 1, 1, 0, 0, 0, 0],
    '3x1x2x2': [1, 1, 1, 0, 1, 1, 0, 0],  # Ритм: 3р 1в 2р 2в
    '1x6': [1, 0, 0, 0, 0, 0, 0]
}


# ================= ЛОГИКА ГРАФИКОВ =================

def normalize_key(s):
    if pd.isna(s): return ""
    s = str(s).lower().replace(' ', '').replace('*', 'x').replace('х', 'x')
    return s


def is_holiday(d):
    return (d.month, d.day) in HOLIDAYS_2026


def get_5x2_val_for_date(d, mode):
    """Логика 5х2 с праздниками"""
    if is_holiday(d) or d.weekday() >= 5:
        return REST
    norm_mode = normalize_key(mode)
    return SHIFT_2 if '2' in norm_mode and '1' not in norm_mode else SHIFT_1


def build_cycle_for_pattern(pattern_name, mode, mask):
    """
    Строит цикл смен с учетом специальной логики для 3x2x3x1 и 3x1x2x2 в режиме 1x2.
    """
    norm_mode = normalize_key(mode)

    # Обработка специальных графиков в режиме 1x2
    if '1x2' in norm_mode and ('3x2x3x1' in pattern_name or '3x1x2x2' in pattern_name):
        # Разбиваем маску на блоки до выходных
        blocks = []
        current_block = []

        for i, val in enumerate(mask):
            if val == 1:
                current_block.append(val)
            else:
                if current_block:
                    blocks.append(current_block)
                    current_block = []
                # Добавляем выходной(ые) как отдельный блок
                blocks.append([0])  # Каждый выходной — отдельный блок

        if current_block:
            blocks.append(current_block)

        # Чередуем смены по блокам: первый блок — 2, второй — 1, третий — 2...
        cycle = []
        shift_toggle = 2  # Начинаем с вечерней (как в примере: 2 2 2 В В 1 1 1 В)

        for block in blocks:
            for _ in block:
                if block[0] == 1:  # Рабочий блок
                    cycle.append(shift_toggle)
                else:  # Выходной
                    cycle.append(REST)
            # Переключаем смену после каждого блока (даже после выходных!)
            shift_toggle = 1 if shift_toggle == 2 else 2

        return cycle

    # Обычная логика для остальных
    if '1x2' in norm_mode:
        part1 = [(SHIFT_1 if x else REST) for x in mask]
        part2 = [(SHIFT_2 if x else REST) for x in mask]
        return part1 + part2
    elif '2' in norm_mode:
        return [(SHIFT_2 if x else REST) for x in mask]
    else:
        return [(SHIFT_1 if x else REST) for x in mask]


def solve_cyclic(feb_vals, pattern_name, mode):
    pkey = normalize_key(pattern_name)
    mask = None
    for k, v in CYCLIC_PATTERNS.items():
        if normalize_key(k) == pkey:
            mask = v
            break
    if not mask: return None

    # Строим цикл с учетом новой логики
    cycle = build_cycle_for_pattern(pattern_name, mode, mask)

    # Поиск смещения
    cycle_len = len(cycle)
    candidates = []

    for offset in range(cycle_len):
        score = 0
        mismatch = False
        for i, val in enumerate(feb_vals):
            theo = cycle[(offset + i) % cycle_len]

            val_s = str(val).replace('.0', '')
            if val_s.lower() in ['b', 'v']: val_s = REST

            is_w_act = (val_s in ['1', '2'])
            is_w_theo = (str(theo) in ['1', '2'])

            if is_w_act != is_w_theo:
                mismatch = True
                break

            if val_s == str(theo): score += 1

        if not mismatch:
            candidates.append((score, offset))

    if not candidates: return None

    best_offset = max(candidates, key=lambda x: x[0])[1]

    # Генерация года
    jan1_offset = (best_offset - 31) % cycle_len
    full_seq = []
    for i in range(365):
        full_seq.append(cycle[(jan1_offset + i) % cycle_len])

    return full_seq


def solve_5x2(feb_vals, mode):
    full_seq = []
    start_date = date(2026, 1, 1)

    for i in range(365):
        curr_date = start_date + timedelta(days=i)
        val = get_5x2_val_for_date(curr_date, mode)
        full_seq.append(val)

    # Простая проверка совпадений (не строгая)
    feb_slice = full_seq[31:59]
    matches = sum(1 for i in range(28) if str(feb_vals[i]).replace('.0', '') == str(feb_slice[i]))
    if matches < 15:
        print("Предупреждение: 5х2 не совпало более чем на 50%")

    return full_seq


# ================= ЭКСПОРТ В EXCEL С ФОРМАТИРОВАНИЕМ =================

def format_excel_file(filepath, month_num):
    """
    Открывает файл, применяет форматирование, пересчитывает 'вых.' и сохраняет.
    """
    wb = load_workbook(filepath)
    ws = wb.active

    # Шрифт Verdana 12 для всего
    font = Font(name='Verdana', size=12)

    # Заливка для выходных
    holiday_fill = PatternFill(start_color="FFB7FD", end_color="FFB7FD", fill_type="solid")

    # Определяем даты месяца
    year = 2026
    _, days_in_month = calendar.monthrange(year, month_num)
    start_date = date(year, month_num, 1)

    # Индекс колонки "вых." (предполагаем, что она есть)
    col_names = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    try:
        col_vyh_index = col_names.index('вых.') + 1  # +1 потому что openpyxl индексирует с 1
    except ValueError:
        print(f"В файле {filepath} не найдена колонка 'вых.'. Пропуск форматирования этой колонки.")
        col_vyh_index = None

    # Проходим по всем строкам (начиная со 2-й, т.к. 1-я — заголовки)
    for row_idx in range(2, ws.max_row + 1):
        # Подсчет выходных для колонки "вых."
        if col_vyh_index:
            rest_count = 0
            for day_col in range(1, days_in_month + 1):
                # Ищем колонку дня (по названию столбца)
                day_col_letter = None
                for c in range(1, ws.max_column + 1):
                    if str(ws.cell(row=1, column=c).value) == str(day_col):
                        day_col_letter = c
                        break
                if day_col_letter:
                    cell_val = str(ws.cell(row=row_idx, column=day_col_letter).value).strip()
                    if cell_val == 'В':
                        rest_count += 1
            ws.cell(row=row_idx, column=col_vyh_index, value=rest_count)

        # Форматирование дней
        for day_col in range(1, days_in_month + 1):
            # Находим колонку дня
            col_letter = None
            for c in range(1, ws.max_column + 1):
                if str(ws.cell(row=1, column=c).value) == str(day_col):
                    col_letter = c
                    break
            if not col_letter:
                continue

            # Применяем шрифт ко всем ячейкам
            cell = ws.cell(row=row_idx, column=col_letter)
            cell.font = font
            cell.alignment = Alignment(horizontal='center')

            # Проверяем, является ли день календарным выходным (для заливки)
            curr_date = start_date.replace(day=day_col)
            if curr_date.weekday() >= 5 or is_holiday(curr_date):
                cell.fill = holiday_fill

    # Автоподбор ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Получаем букву столбца
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)  # Ограничиваем макс. ширину
        ws.column_dimensions[column].width = adjusted_width

    wb.save(filepath)


# ================= MAIN =================

def main():
    print("--- Генератор табелей v5 (Специальные графики + Форматирование) ---")

    if not os.path.exists(INPUT_FILE):
        print("Файл не найден")
        return

    df = pd.read_excel(INPUT_FILE, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]

    day_cols = []
    for i in range(1, 29):
        if str(i) in df.columns: day_cols.append(str(i))

    meta_cols = ['Таб.№', 'График', 'Режим', 'см.', 'вых.']
    for c in meta_cols:
        if c not in df.columns: df[c] = ""

    full_year_map = {}
    stats_ok = 0

    for idx, row in df.iterrows():
        grafik = normalize_key(row['График'])
        mode = str(row['Режим'])
        feb_vals = []
        for d in day_cols:
            val = str(row[d]).strip()
            if val.lower() in ['b', 'v', 'nan']: val = REST
            feb_vals.append(val)

        result_seq = None
        if '5x2' in grafik:
            result_seq = solve_5x2(feb_vals, mode)
        else:
            result_seq = solve_cyclic(feb_vals, grafik, mode)

        if result_seq:
            full_year_map[idx] = result_seq
            stats_ok += 1
        else:
            print(f"Строка {idx}: Не удалось построить график {grafik}")
            full_year_map[idx] = [""] * 365

    print(f"Успешно обработано: {stats_ok}")

    if not os.path.exists(OUTPUT_DIR): os.makedirs(OUTPUT_DIR)

    months = [1] + list(range(3, 13))
    m_names = {1: 'january', 2: 'february', 3: 'march', 4: 'april', 5: 'may', 6: 'june',
               7: 'july', 8: 'august', 9: 'september', 10: 'october', 11: 'november', 12: 'december'}

    for m in months:
        fname = f"{m_names[m]}_2026.xlsx"
        print(f"Создаем {fname}...")

        _, days_cnt = calendar.monthrange(2026, m)
        doy_start = date(2026, m, 1).timetuple().tm_yday - 1

        new_df = df[meta_cols].copy()

        for d in range(1, days_cnt + 1):
            vals = []
            d_idx = doy_start + (d - 1)
            for idx in df.index:
                vals.append(full_year_map[idx][d_idx])
            new_df[str(d)] = vals

        out_path = os.path.join(OUTPUT_DIR, fname)
        new_df.to_excel(out_path, index=False)

        # Применяем форматирование
        format_excel_file(out_path, m)
        print(f" -> Форматирование применено.")

    print("✅ Все готово!")


if __name__ == "__main__":
    main()