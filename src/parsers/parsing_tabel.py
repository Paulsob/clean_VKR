import json
import calendar
from pathlib import Path
from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent

TABELS_DIR = PROJECT_ROOT / "data" / "tabeles_2026"
OUTPUT_DIR = PROJECT_ROOT / "data" / "drivers_json"
YEAR = 2026

# Английские → русские названия месяцев (для содержимого JSON)
EN_TO_RU_MONTH = {
    "january": "Январь",
    "february": "Февраль",
    "march": "Март",
    "april": "Апрель",
    "may": "Май",
    "june": "Июнь",
    "july": "Июль",
    "august": "Август",
    "september": "Сентябрь",
    "october": "Октябрь",
    "november": "Ноябрь",
    "december": "Декабрь",
}

# Для определения количества дней в месяце
EN_MONTH_TO_NUM = {en: i + 1 for i, en in enumerate([
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december"
])}

MONTH_ORDER = list(EN_TO_RU_MONTH.keys())
SHEET_NAME = "Sheet1"


def extract_english_month(filename: str) -> str | None:
    """Извлекает английское название месяца из имени файла (например, 'january_2026.xlsx' → 'january')."""
    stem = filename.lower().replace(".xlsx", "")
    for en_month in MONTH_ORDER:
        if en_month in stem:
            return en_month
    return None


def parse_whole_sheet(sheet_rows, en_month: str, year: int):
    """Парсит одну таблицу в структуру водителей."""
    month_num = EN_MONTH_TO_NUM[en_month]
    days_in_month = calendar.monthrange(year, month_num)[1]

    if not sheet_rows:
        return []

    header = sheet_rows[0]
    total_columns_expected = len(header)
    drivers = []

    for row in sheet_rows[1:]:
        if len(row) < total_columns_expected:
            row = list(row) + [None] * (total_columns_expected - len(row))
        if len(row) < 5:
            continue

        try:
            tab_number = int(row[0])
        except (ValueError, TypeError):
            continue

        schedule = str(row[1]).strip() if row[1] is not None else ""
        mode = str(row[2]).strip() if row[2] is not None else ""

        day_values = row[5:]
        days = []
        for day in range(1, days_in_month + 1):
            idx = day - 1
            value = day_values[idx] if idx < len(day_values) else None
            value_str = str(value).strip() if value is not None else ""
            days.append({"day": day, "value": value_str})

        drivers.append({
            "tab_number": tab_number,
            "schedule": schedule,
            "mode": mode,
            "days": days
        })

    return drivers


def main():
    if not TABELS_DIR.exists():
        raise FileNotFoundError(f"Папка с табелями не найдена: {TABELS_DIR}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Сортируем файлы по порядку месяцев
    files = sorted(
        TABELS_DIR.glob("*.xlsx"),
        key=lambda f: next(
            (i for i, m in enumerate(MONTH_ORDER) if m in f.name.lower()), 999
        )
    )

    processed = 0

    for file_path in files:
        en_month = extract_english_month(file_path.name)
        if not en_month:
            print(f"⚠️ Пропущен файл (не распознан месяц): {file_path.name}")
            continue

        ru_month = EN_TO_RU_MONTH[en_month]
        print(f"Обрабатываю: {file_path.name} → {ru_month} {YEAR}")

        try:
            workbook = load_workbook(file_path, data_only=True)
            if SHEET_NAME not in workbook.sheetnames:
                print(f"  ❌ Лист '{SHEET_NAME}' не найден в {file_path.name}")
                continue

            sheet = workbook[SHEET_NAME]
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            drivers = parse_whole_sheet(rows, en_month, YEAR)

            month_num = EN_MONTH_TO_NUM[en_month]

            # Имя файла: drivers_january.json
            output_file = OUTPUT_DIR / f"{month_num:02d}_drivers_{en_month}.json"
            result_data = {
                "month": ru_month,
                "year": YEAR,
                "drivers": drivers
            }

            with open(output_file, "w", encoding="utf-8") as f:
                json.dump(result_data, f, ensure_ascii=False, indent=2)

            processed += 1

        except Exception as e:
            print(f"  ❌ Ошибка при обработке {file_path.name}: {e}")

    print(f"\n✅ Успешно обработано {processed} месяцев.")
    print(f"Файлы сохранены в: {OUTPUT_DIR.absolute()}")


if __name__ == "__main__":
    main()