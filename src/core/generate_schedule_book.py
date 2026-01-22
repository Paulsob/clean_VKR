import sys
import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font

# === НАСТРОЙКИ ПУТЕЙ ===
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(current_dir))
sys.path.insert(0, project_root)
os.chdir(project_root)

from src.config import (
    SELECTED_ROUTE, SELECTED_MONTH, SELECTED_YEAR,
    SIMULATION_RESULT_FILE, OUTPUTS_DIR
)

OUTPUT_FILE = os.path.join(OUTPUTS_DIR, f"_schedule_book_{SELECTED_ROUTE}_{SELECTED_MONTH}.xlsx")


def main():
    print(f"--- ГЕНЕРАЦИЯ ЖУРНАЛА НАРЯДОВ ---")

    if not os.path.exists(SIMULATION_RESULT_FILE):
        print("Файл симуляции не найден. Запустите run_simulation.py")
        return

    with open(SIMULATION_RESULT_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Сортируем дни (1, 2, 3...)
    days_sorted = sorted([k for k in data.keys() if k.isdigit()], key=int)

    rows = []

    print("Формирование строк...")
    for day_str in days_sorted:
        day_res = data[day_str]
        roster = day_res.get("roster", [])

        # Сортируем вагоны по номеру (как числа)
        try:
            roster.sort(key=lambda x: int(x.get("tram_number", 0)))
        except:
            pass

        for tram in roster:
            t_num = tram.get("tram_number", "?")

            # Данные 1 смены
            s1 = tram.get("shift_1") or {}
            d1 = s1.get("driver", "---")
            t1 = s1.get("time_range", "??:??-??:??")
            w1 = "(!)" if s1.get("warnings") else ""

            # Данные 2 смены
            s2 = tram.get("shift_2") or {}
            d2 = s2.get("driver", "---")
            t2 = s2.get("time_range", "??:??-??:??")
            w2 = "(!)" if s2.get("warnings") else ""

            # Формируем строку
            row = {
                "Дата": int(day_str),
                "День": f"{day_str} {SELECTED_MONTH}",
                "Вагон": int(t_num) if str(t_num).isdigit() else t_num,

                "I Смена (Водитель)": f"{d1} {w1}".strip(),
                "I Время": t1,

                "II Смена (Водитель)": f"{d2} {w2}".strip(),
                "II Время": t2,

                "Проблемы": ", ".join(tram.get("issues", []))
            }
            rows.append(row)

    # Создаем DataFrame
    df = pd.DataFrame(rows)

    # Сохраняем
    os.makedirs(OUTPUTS_DIR, exist_ok=True)
    writer = pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name="Журнал")

    # Оформление
    ws = writer.book.active

    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    fill_empty = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Красный если пусто
    fill_warn = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Желтый если нарушение

    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # Настраиваем заголовки
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Проходим по данным
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            val = str(cell.value)

            # Подсветка проблем
            if "---" in val:
                cell.fill = fill_empty
            elif "(!)" in val:
                cell.fill = fill_warn

    # Ширина колонок
    ws.column_dimensions['A'].width = 5  # Дата (число)
    ws.column_dimensions['B'].width = 15  # Дата (текст)
    ws.column_dimensions['C'].width = 8  # Вагон
    ws.column_dimensions['D'].width = 20  # Водитель 1
    ws.column_dimensions['E'].width = 15  # Время 1
    ws.column_dimensions['F'].width = 20  # Водитель 2
    ws.column_dimensions['G'].width = 15  # Время 2
    ws.column_dimensions['H'].width = 30  # Проблемы

    writer.close()
    print(f"✅ Детальный журнал создан: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()