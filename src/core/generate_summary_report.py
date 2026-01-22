import sys
import os
import json
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# === НАСТРОЙКА ПУТЕЙ ===
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(current_dir))
sys.path.insert(0, project_root)
os.chdir(project_root)

from src.config import (
    SELECTED_ROUTE, SELECTED_MONTH, SELECTED_YEAR,
    SIMULATION_RESULT_FILE, REPORT_FILE
)
from src.database import DataLoader


def load_absences():
    """Загружает больничные и отпуска"""
    absences_path = "data/absences.json"
    absences = {}  # { driver_id: { date: type } }

    if not os.path.exists(absences_path):
        return absences

    with open(absences_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    for item in data.get("absences", []):
        did = str(item["driver_id"])
        d_from = datetime.strptime(item["from"], "%Y-%m-%d").date()
        d_to = datetime.strptime(item["to"], "%Y-%m-%d").date()
        a_type = item["type"]  # 'sick' или 'vacation'

        if did not in absences:
            absences[did] = {}

        # Заполняем все дни диапазона
        current = d_from
        while current <= d_to:
            absences[did][current] = a_type
            current += pd.Timedelta(days=1).to_pytimedelta()
            # Или: current = current + timedelta(days=1)

    return absences


def main():
    print(f"--- ГЕНЕРАЦИЯ СВОДНОГО ОТЧЕТА ---")

    # 1. Загрузка данных
    if not os.path.exists(SIMULATION_RESULT_FILE):
        print("Файл симуляции не найден!")
        return

    with open(SIMULATION_RESULT_FILE, "r", encoding="utf-8") as f:
        sim_data = json.load(f)

    db = DataLoader()
    db.load_all()

    absences = load_absences()

    # Водители маршрута
    drivers_db = {
        str(d.id): d for d in db.drivers
        if str(d.assigned_route_number) == str(SELECTED_ROUTE) and d.month == SELECTED_MONTH
    }

    if not drivers_db:
        print("Водители не найдены!")
        return

    # Карта месяцев
    month_map = {"Январь": 1, "Февраль": 2, "Март": 3, "Апрель": 4, "Май": 5, "Июнь": 6,
                 "Июль": 7, "Август": 8, "Сентябрь": 9, "Октябрь": 10, "Ноябрь": 11, "Декабрь": 12}
    m_num = month_map.get(SELECTED_MONTH, 2)

    all_days = sorted([int(k) for k in sim_data.keys() if k.isdigit()])

    # 2. Сбор данных
    # { driver_id: { 'work_count': 0, 'reserve_count': 0, 'total_hours': 0, 'days': {...} } }
    report_data = {}

    for did in drivers_db.keys():
        report_data[did] = {
            'work_count': 0,
            'reserve_count': 0,
            'total_hours': 0.0,
            'days': {}
        }

    # Заполняем ФАКТ (кто работал)
    for day_str, day_res in sim_data.items():
        if not day_str.isdigit(): continue
        day = int(day_str)

        for tram in day_res.get("roster", []):
            for shift_key in ["shift_1", "shift_2"]:
                s_info = tram.get(shift_key)
                if s_info and s_info.get("driver"):
                    did = s_info["driver"].split(" ")[0]
                    if did in report_data:
                        wh = s_info.get("work_hours", 8.0)
                        rest = s_info.get("rest_before", 0)

                        rest_str = f"{rest:.0f}" if rest < 100 else ">48"
                        val = f"{wh:.1f}ч (отд {rest_str})"
                        if s_info.get("warnings"): val += " (!)"

                        report_data[did]['days'][day] = val
                        report_data[did]['work_count'] += 1
                        report_data[did]['total_hours'] += wh

    # Заполняем ОСТАЛЬНЫХ (Резерв, Выходные, Больничные)
    for did, driver_obj in drivers_db.items():
        driver_absences = absences.get(did, {})

        for day in all_days:
            if day in report_data[did]['days']:
                continue  # Уже работал

            # Проверяем больничный/отпуск
            check_date = datetime(SELECTED_YEAR, m_num, day).date()

            if check_date in driver_absences:
                a_type = driver_absences[check_date]
                if a_type == "sick":
                    report_data[did]['days'][day] = "БОЛЬНИЧНЫЙ"
                else:
                    report_data[did]['days'][day] = "ОТПУСК"
                continue

            # Проверяем табель
            plan = driver_obj.get_status_for_day(day)

            if plan in ["1", "2"]:
                report_data[did]['days'][day] = "РЕЗЕРВ"
                report_data[did]['reserve_count'] += 1
            else:
                report_data[did]['days'][day] = plan  # "В", "О", "Б" и т.д.

    # 3. Разделение на группы
    # Активные = те, кто хотя бы раз работал
    # Резерв = те, кто НИ РАЗУ не работал (только резерв или выходные)

    active_ids = []
    reserve_ids = []

    for did, data in report_data.items():
        if data['work_count'] > 0:
            active_ids.append(did)
        else:
            reserve_ids.append(did)

    # Сортировка
    active_ids.sort(key=lambda x: int(x) if x.isdigit() else x)
    reserve_ids.sort(key=lambda x: int(x) if x.isdigit() else x)

    # 4. Подсчет статистики
    total_shifts_needed = 0
    for day_str, day_res in sim_data.items():
        if not day_str.isdigit(): continue
        for tram in day_res.get("roster", []):
            if tram.get("shift_1"): total_shifts_needed += 1
            if tram.get("shift_2"): total_shifts_needed += 1

    total_active = len(active_ids)
    total_reserve = len(reserve_ids)
    total_all = total_active + total_reserve

    # 5. Формирование DataFrame
    def make_row(did):
        data = report_data[did]
        row = {
            "Таб.№": did,
            "Смен": data['work_count'],
            "Часов": round(data['total_hours'], 1),
            "Резерв (дн)": data['reserve_count']
        }
        for day in all_days:
            row[day] = data['days'].get(day, "")
        return row

    # Строим таблицу
    rows = []

    # Заголовок группы "АКТИВНЫЕ"
    header_active = {"Таб.№": "═══ РАБОЧЕЕ ЯДРО ═══", "Смен": "", "Часов": "", "Резерв (дн)": ""}
    for day in all_days: header_active[day] = ""
    rows.append(header_active)

    for did in active_ids:
        rows.append(make_row(did))

    # Разделитель
    separator = {"Таб.№": "", "Смен": "", "Часов": "", "Резерв (дн)": ""}
    for day in all_days: separator[day] = ""
    rows.append(separator)

    # Заголовок группы "РЕЗЕРВ"
    header_reserve = {"Таб.№": "═══ ИЗБЫТОЧНЫЙ РЕЗЕРВ ═══", "Смен": "", "Часов": "", "Резерв (дн)": ""}
    for day in all_days: header_reserve[day] = ""
    rows.append(header_reserve)

    for did in reserve_ids:
        rows.append(make_row(did))

    # Статистика
    rows.append(separator)
    stat1 = {"Таб.№": f"Всего смен в месяц: {total_shifts_needed}", "Смен": "", "Часов": "", "Резерв (дн)": ""}
    # stat2 = {"Таб.№": f"Активных водителей: {total_active}", "Смен": "", "Часов": "", "Резерв (дн)": ""}
    # stat3 = {"Таб.№": f"В избытке: {total_reserve}", "Смен": "", "Часов": "", "Резерв (дн)": ""}
    for day in all_days:
        stat1[day] = ""
        # stat2[day] = ""
        # stat3[day] = ""
    rows.append(stat1)
    # rows.append(stat2)
    # rows.append(stat3)

    df = pd.DataFrame(rows)

    # 6. Экспорт в Excel
    os.makedirs(os.path.dirname(REPORT_FILE), exist_ok=True)
    writer = pd.ExcelWriter(REPORT_FILE, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name="Сводный отчет")

    # Оформление
    ws = writer.book.active

    # Цвета
    fill_work = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Зеленый
    fill_work_warn = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Оранжевый (нарушение)
    fill_reserve = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Красный (резерв)
    fill_rest = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Серый (выходной)
    fill_sick = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Желтый (больничный/отпуск)
    fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Синий (заголовки групп)

    font_header = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        first_cell_val = str(row[0].value) if row[0].value else ""

        # Заголовки групп
        if "═══" in first_cell_val:
            for cell in row:
                cell.fill = fill_header
                cell.font = font_header
            continue

        # Обычные строки
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            val = str(cell.value) if cell.value else ""

            if "БОЛЬНИЧНЫЙ" in val or "ОТПУСК" in val:
                cell.fill = fill_sick
            elif "(!)" in val:
                cell.fill = fill_work_warn
            elif "ч (" in val:
                cell.fill = fill_work
            elif "РЕЗЕРВ" in val:
                cell.fill = fill_reserve
            elif val in ["В", "B", "О", "Б"]:
                cell.fill = fill_rest

    # Ширина колонок
    ws.column_dimensions['A'].width = 30  # Для статистики
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 12
    for col in range(5, len(all_days) + 5):
        col_letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[col_letter].width = 14

    writer.close()

    # Вывод статистики в консоль
    print("\n" + "=" * 50)
    print(f"СТАТИСТИКА МАРШРУТА {SELECTED_ROUTE}")
    print("=" * 50)
    print(f"Всего смен в месяц: {total_shifts_needed}")
    print(f"Активных водителей (реально работали): {total_active}")
    print(f"В избытке (ни разу не вышли): {total_reserve}")
    print(f"Всего по табелю: {total_all}")
    print("=" * 50)
    print(f"✅ Отчет сохранен: {REPORT_FILE}")


if __name__ == "__main__":
    main()